"""Order battles to minimize per-member character switches, and write output xlsx."""
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import SCHEDULE_SHEET, SUMMARY_SHEET, TARGETS, TARGET_TO_MAP
from .optimize import Battle


def drop_zero_credit_battles(
    battles: List[Battle], quests: Dict
) -> List[Battle]:
    """Remove battles that contribute no NEW quest completion.

    Uses a greedy set-cover pass: repeatedly keeps the battle that credits the
    most so-far-uncredited (member, character, target) triples, until every
    quest-credit achievable from the input list has been covered. Battles that
    would only re-credit an already-credited quest are dropped.
    """
    if not battles:
        return []

    def provides(b: Battle) -> set:
        return {
            (m, c, b.target)
            for m, c in b.participants.items()
            if quests.get((m, c, b.target), 0) == 1
        }

    remaining = list(battles)
    covered: set = set()
    kept: List[Battle] = []

    while True:
        best: Optional[Battle] = None
        best_new: set = set()
        for b in remaining:
            new = provides(b) - covered
            if len(new) > len(best_new):
                best = b
                best_new = new
        if best is None or not best_new:
            break
        kept.append(best)
        covered |= best_new
        remaining.remove(best)

    # Preserve the solver's original ordering among kept battles.
    order_index = {id(b): i for i, b in enumerate(battles)}
    kept.sort(key=lambda b: order_index[id(b)])
    return kept


def _switch_cost(prev: Battle | None, b: Battle) -> int:
    """Number of member-level character switches when going prev -> b.

    Wildcards (participant value == None) are treated as "will be filled to
    match whatever is needed", so they never count as a switch.
    """
    if prev is None:
        return 0
    cost = 0
    for m, c in b.participants.items():
        if c is None:
            continue
        pc = prev.participants.get(m)
        if pc is None or pc == c:
            continue
        cost += 1
    return cost


# Penalty added when every shared member switches characters (full re-team).
# Full re-team forces a trip through campus to re-team, so this is the
# heaviest per-transition cost.
_FULL_RETEAM_PENALTY = 1000

# Penalty for changing maps (targets in different map groups). Cheaper than a
# full re-team but more expensive than a single character switch, so the
# heuristics will cluster same-map battles while still allowing a map change
# to save multiple switches if worthwhile.
_MAP_CHANGE_PENALTY = 10


def _transition_score(prev: Battle | None, b: Battle) -> int:
    """Switch cost with penalties for map change and full re-team.

    Costs (additive):
      * +1 per member who participates in both and switches character
        (wildcards / None are ignored).
      * +_MAP_CHANGE_PENALTY if the two battles are in different map groups.
      * +_FULL_RETEAM_PENALTY if every shared concrete member switches
        (forces a return to campus to re-team).
    """
    if prev is None:
        return 0
    switches = 0
    shared = 0
    for m, c in b.participants.items():
        if c is None:
            continue
        pc = prev.participants.get(m)
        if pc is None:
            continue
        shared += 1
        if pc != c:
            switches += 1
    score = switches
    if TARGET_TO_MAP.get(prev.target) != TARGET_TO_MAP.get(b.target):
        score += _MAP_CHANGE_PENALTY
    if shared > 0 and switches == shared:
        score += _FULL_RETEAM_PENALTY
    return score


def _solidify_wildcards(
    battles: List[Battle], members: List[str], quests: Dict
) -> None:
    """In-place: set participants[m] = None for any slot that is neither the
    battle's ticket source nor a fresh (first-seen) quest-credit for that
    member at that target.

    After this call, every concrete char in a battle is something the plan
    _requires_ to be exactly that character; all other slots are wildcards
    that the ordering heuristic can move freely.
    """
    for m in members:
        covered: set = set()  # (char, target) already used as a credit anchor

        # Pass A: ticket sources — their character is locked; also counts as
        # coverage if the locked char credits a quest.
        for b in battles:
            if m not in b.participants:
                continue
            if b.ticket_source[0] == m:
                c = b.ticket_source[1]
                b.participants[m] = c
                if quests.get((m, c, b.target), 0) == 1:
                    covered.add((c, b.target))

        # Pass B: for non-source slots, keep the char only if it brings a
        # still-uncovered quest credit.
        for b in battles:
            if m not in b.participants:
                continue
            if b.ticket_source[0] == m:
                continue
            c = b.participants[m]
            if c is None:
                continue
            key = (c, b.target)
            if quests.get((m, c, b.target), 0) == 1 and key not in covered:
                covered.add(key)
                # keep concrete
            else:
                b.participants[m] = None  # wildcard


def _fill_wildcards(
    battles: List[Battle],
    members: List[str],
    chars_by_member: Dict[str, List[str]],
    quests: Dict | None = None,
) -> None:
    """Replace every wildcard participant slot (None) with a concrete char
    from the member's roster, chosen to minimize switches.

    Strategy per member: walk their participations in final order.
      1. If previous concrete char exists and is in roster, reuse it
         (no switch). This is the dominant case.
      2. Else, peek ahead to the next concrete char and reuse it
         (saves the upcoming forced switch).
      3. Else, if `quests` is provided, prefer a char that credits a quest
         at this target (may enable later battles to be dropped).
      4. Fall back to roster[0].
    """
    for m in members:
        roster = list(chars_by_member.get(m, []))
        if not roster:
            continue

        idx = [i for i, b in enumerate(battles) if m in b.participants]
        chars: List[Optional[str]] = [battles[i].participants[m] for i in idx]
        targets = [battles[i].target for i in idx]

        prev_c: Optional[str] = None
        for k in range(len(chars)):
            if chars[k] is None:
                choice: Optional[str] = None
                if prev_c is not None:
                    choice = prev_c
                else:
                    # No previous concrete char — peek ahead.
                    choice = next(
                        (chars[j] for j in range(k + 1, len(chars))
                         if chars[j] is not None),
                        None,
                    )
                    if choice is None and quests is not None:
                        # Prefer a quest-crediting char at this target.
                        t = targets[k]
                        choice = next(
                            (c for c in roster
                             if quests.get((m, c, t), 0) == 1),
                            None,
                        )
                    if choice is None:
                        choice = roster[0]
                chars[k] = choice
            prev_c = chars[k]

        for k, i in enumerate(idx):
            battles[i].participants[m] = chars[k]  # type: ignore[assignment]


def _break_full_reteams(battles: List[Battle]) -> List[Battle]:
    """Local pairwise swap pass that removes full-reteam transitions.

    Looks at each adjacent boundary (i, i+1). If it is a full re-team, tries
    swapping battle i+1 with any later battle j; accepts the first swap that
    strictly reduces the total transition score (i.e. breaks the full re-team
    without creating another one and without worsening overall switches).
    Runs until no beneficial swap is found, capped at a few sweeps.
    """
    if len(battles) < 2:
        return battles

    def total_score(bs: List[Battle]) -> int:
        return sum(_transition_score(bs[i - 1], bs[i]) for i in range(1, len(bs)))

    for _ in range(5):  # a handful of sweeps is enough in practice
        improved = False
        i = 1
        while i < len(battles):
            # Detect full re-team at boundary (i-1, i)
            prev, cur = battles[i - 1], battles[i]
            shared = sum(1 for m in cur.participants if m in prev.participants)
            if shared > 0 and _switch_cost(prev, cur) == shared:
                base = total_score(battles)
                best_j = None
                best_score = base
                for j in range(i + 1, len(battles)):
                    battles[i], battles[j] = battles[j], battles[i]
                    s = total_score(battles)
                    if s < best_score:
                        best_score = s
                        best_j = j
                    battles[i], battles[j] = battles[j], battles[i]
                if best_j is not None:
                    battles[i], battles[best_j] = battles[best_j], battles[i]
                    improved = True
            i += 1
        if not improved:
            break
    return battles


def order_battles(battles: List[Battle]) -> List[Battle]:
    """Greedy nearest-neighbor ordering minimizing switches.

    Tries each battle as a starting point and keeps the best sequence.
    Works well for the problem size (usually < 50 battles).
    """
    if not battles:
        return []

    best_seq: List[Battle] = []
    best_cost = None

    for start_idx in range(len(battles)):
        remaining = battles.copy()
        seq = [remaining.pop(start_idx)]
        total = 0
        while remaining:
            prev = seq[-1]
            # Pick remaining battle with minimum transition score. Tie-break:
            # same target > same map > different map.
            def key(b: Battle):
                same_target = 0 if b.target == prev.target else 1
                same_map = 0 if TARGET_TO_MAP.get(b.target) == TARGET_TO_MAP.get(prev.target) else 1
                return (_transition_score(prev, b), same_map, same_target)
            nxt = min(remaining, key=key)
            total += _transition_score(prev, nxt)
            remaining.remove(nxt)
            seq.append(nxt)
        if best_cost is None or total < best_cost:
            best_cost = total
            best_seq = seq

    return _break_full_reteams(best_seq)


def finalize_schedule(
    battles: List[Battle],
    members: List[str],
    quests: Dict,
    chars_by_member: Dict[str, List[str]] | None = None,
) -> List[Battle]:
    """Drop zero-credit battles, order, and reduce switches to a fixpoint.

    Iterates drop-zero-credit -> order -> reduce-switches until the pipeline
    is stable: battle count does not change AND no battle ends with zero
    credited quests. Bounded to 50 iterations as a safety net.

    When `chars_by_member` is provided, filler (non-ticket-source) slots may
    be assigned any character from the member's full roster, which allows
    far more aggressive switch reduction than multiset-preserving reassignment.
    """
    def _zero_count(bs: List[Battle]) -> int:
        credited: set = set()
        zero = 0
        for b in bs:
            new = {(m, c, b.target)
                   for m, c in b.participants.items()
                   if quests.get((m, c, b.target), 0) == 1} - credited
            if not new:
                zero += 1
            credited |= new
        return zero

    battles = list(battles)
    prev_n = -1
    for _ in range(50):
        battles = drop_zero_credit_battles(battles, quests)
        if chars_by_member is not None:
            # Wildcard flow: erase non-forced chars, order with wildcard
            # awareness, then fill wildcards to match neighbors.
            _solidify_wildcards(battles, members, quests)
            battles = order_battles(battles)
            _fill_wildcards(battles, members, chars_by_member, quests)
            # Filled wildcards may coincidentally credit extra quests, which
            # can make other battles redundant — drop again on filled state.
            battles = drop_zero_credit_battles(battles, quests)
        else:
            battles = order_battles(battles)
            reduce_switches(battles, members, quests, chars_by_member)
        battles = _break_full_reteams(battles)
        if len(battles) == prev_n and _zero_count(battles) == 0:
            break
        prev_n = len(battles)

    # One last defensive filter: drop any residual zero-credit battle.
    # (drop_zero_credit_battles already does this; cheap to double-check.)
    battles = drop_zero_credit_battles(battles, quests)
    # Final polish: re-solidify, re-order, re-fill so ordering uses only
    # required chars and fillers match their neighbors.
    if chars_by_member is not None:
        _solidify_wildcards(battles, members, quests)
        battles = order_battles(battles)
        _fill_wildcards(battles, members, chars_by_member, quests)
    else:
        battles = order_battles(battles)
    return battles


def reduce_switches(
    battles: List[Battle],
    members: List[str],
    quests: Dict,
    chars_by_member: Dict[str, List[str]] | None = None,
) -> None:
    """In-place: reassign filler participant slots to minimize switches.

    A participant slot is 'fixed' when the member is the battle's ticket
    source — that character is forced by the ticket. All other slots are
    filler: the MILP does NOT constrain which character of the member
    participates (fillers consume no ticket). Therefore any character from
    the member's full roster (`chars_by_member[m]`) is a legal filler.

    This function preserves:
      * Which battles each member joins (participation pattern).
      * Ticket-source assignments (character locked for that member).
      * All currently-credited weekly quests — for each (member, target),
        every character that currently credits a quest is guaranteed to
        appear in at least one target-t battle the member joins.

    When `chars_by_member` is None, the earlier multiset-preserving behavior
    is used as a safe fallback (less effective at reducing switches).
    """
    if chars_by_member is None:
        _reduce_switches_multiset(battles, members, quests)
        return

    for m in members:
        roster = list(chars_by_member.get(m, []))
        if not roster:
            continue

        # For each target, collect the chars that currently credit a quest
        # for this member (we must preserve at least one placement each).
        current_credits: Dict[str, set] = defaultdict(set)
        for b in battles:
            if m not in b.participants:
                continue
            c = b.participants[m]
            if quests.get((m, c, b.target), 0) == 1:
                current_credits[b.target].add(c)

        # Chars forced by this member being the ticket source at a target-t battle.
        # These cover some of the pending credits automatically.
        forced_chars_at: Dict[str, set] = defaultdict(set)
        filler_count: Dict[str, int] = defaultdict(int)
        for b in battles:
            if m not in b.participants:
                continue
            if b.ticket_source[0] == m:
                forced_chars_at[b.target].add(b.ticket_source[1])
            else:
                filler_count[b.target] += 1

        # pending[t] = credit chars still awaiting at least one placement.
        pending: Dict[str, set] = defaultdict(set)
        for t, cs in current_credits.items():
            pending[t] = set(cs) - forced_chars_at[t]

        # Chars the member HAS an uncompleted weekly quest for at target t
        # (used to opportunistically gain extra credits at otherwise-free slots).
        quest_chars_at: Dict[str, set] = defaultdict(set)
        for c in roster:
            for t in TARGETS:
                if quests.get((m, c, t), 0) == 1:
                    quest_chars_at[t].add(c)

        prev_char: Optional[str] = None
        placed_at: Dict[str, set] = defaultdict(set)

        for b in battles:
            if m not in b.participants:
                continue
            t = b.target

            if b.ticket_source[0] == m:
                chosen = b.ticket_source[1]
            else:
                pending_t = pending[t]
                remaining_filler = filler_count[t]
                # If this slot is needed to place a still-pending credit, do so.
                must_place_pending = len(pending_t) >= remaining_filler

                if must_place_pending and pending_t:
                    if prev_char in pending_t:
                        chosen = prev_char  # type: ignore[assignment]
                    else:
                        chosen = next(iter(pending_t))
                else:
                    # Slack available. Prefer prev_char (no switch).
                    if prev_char is not None and prev_char in roster:
                        chosen = prev_char
                    else:
                        # No usable prev — opportunistically pick an uncredited
                        # quest char at this target; else roster[0].
                        fresh_quest = next(
                            (c for c in quest_chars_at[t]
                             if c not in placed_at[t]
                             and c not in current_credits[t]),
                            None,
                        )
                        chosen = fresh_quest if fresh_quest is not None else roster[0]

                filler_count[t] -= 1

            b.participants[m] = chosen
            pending[t].discard(chosen)
            placed_at[t].add(chosen)
            prev_char = chosen


def _reduce_switches_multiset(
    battles: List[Battle], members: List[str], quests: Dict
) -> None:
    """Legacy multiset-preserving reassignment (fallback)."""
    for m in members:
        # Group slots by target, split into fixed / free.
        # pool[t] holds characters still available for free assignment at target t.
        pool: Dict[str, Counter] = defaultdict(Counter)
        is_fixed: Dict[int, Optional[str]] = {}

        for i, b in enumerate(battles):
            if m not in b.participants:
                continue
            pool[b.target][b.participants[m]] += 1
            if b.ticket_source[0] == m:
                is_fixed[i] = b.ticket_source[1]
            else:
                is_fixed[i] = None

        # Remove the fixed characters from their target pool (they are locked).
        for i, b in enumerate(battles):
            if m not in b.participants or is_fixed[i] is None:
                continue
            pool[b.target][is_fixed[i]] -= 1
            if pool[b.target][is_fixed[i]] <= 0:
                del pool[b.target][is_fixed[i]]

        # Prefer chars that CREDIT a quest when reusing previous; this matters
        # when multiple characters are available at the same target.
        prev_char: Optional[str] = None
        credited_at: Dict[tuple, bool] = {}     # (m, c, t) -> already placed once

        for i, b in enumerate(battles):
            if m not in b.participants:
                continue

            fixed = is_fixed[i]
            if fixed is not None:
                b.participants[m] = fixed
                credited_at[(m, fixed, b.target)] = True
                prev_char = fixed
                continue

            p = pool[b.target]
            if not p:
                # Shouldn't happen by construction; guard anyway.
                prev_char = b.participants[m]
                continue

            # Preference order: prev_char (reuse) -> any quest-completing char
            # not yet credited -> any char.
            candidates: List[str] = []
            if prev_char is not None and prev_char in p:
                candidates.append(prev_char)
            for c in list(p):
                if c in candidates:
                    continue
                if quests.get((m, c, b.target), 0) == 1 and not credited_at.get((m, c, b.target)):
                    candidates.append(c)
            for c in list(p):
                if c not in candidates:
                    candidates.append(c)

            chosen = candidates[0]
            b.participants[m] = chosen
            p[chosen] -= 1
            if p[chosen] <= 0:
                del p[chosen]
            credited_at[(m, chosen, b.target)] = True
            prev_char = chosen


# ---------- Excel output ----------

HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
HEADER_FONT = Font(bold=True)
SWITCH_FILL = PatternFill("solid", fgColor="FFFCE4D6")  # highlight switches
BOLD_FONT = Font(bold=True)
TARGET_GROUP_FILL_A = PatternFill("solid", fgColor="FFF2F2F2")  # light grey
TARGET_GROUP_FILL_B = PatternFill("solid", fgColor="FFFFE6EA")  # light pink



def write_schedule(
    battles: List[Battle],
    members: List[str],
    out_path: str | Path,
    quests: Dict | None = None,
) -> None:
    """Write the schedule workbook.

    If `quests` is provided, per-battle and per-member quest counts count each
    (member, character, target) quest at most once (the first battle that
    includes that character at that target credits the quest). If `quests` is
    None, falls back to `Battle.completed`, which may double-count when a
    character joins multiple battles of the same target.
    """
    out_path = Path(out_path)
    wb = Workbook()

    # Pre-compute per-battle "new completions", per-member quest tallies, and
    # which (battle, member) cells should be bolded (the participant whose quest
    # is actually credited at that battle).
    credited: set = set()                                 # (m, c, t) already counted
    battle_completed: list[int] = []                      # parallel to `battles`
    bold_members: list[set] = []                          # per-battle set of members to bold
    per_member_q: Dict[str, int] = {m: 0 for m in members}
    for b in battles:
        n = 0
        bm: set = set()
        for m, c in b.participants.items():
            key = (m, c, b.target)
            if quests is not None:
                if quests.get(key, 0) == 1 and key not in credited:
                    credited.add(key)
                    n += 1
                    bm.add(m)
                    per_member_q[m] += 1
            # else: leave n as 0 here; we use b.completed below for display
        battle_completed.append(n if quests is not None else b.completed)
        bold_members.append(bm)

    # --- Schedule sheet ---
    ws = wb.active
    ws.title = SCHEDULE_SHEET
    headers = ["order", "target", "ticket_kind", "ticket_source"] + members + ["quests_completed"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B2"

    prev: Battle | None = None
    for i, (b, bc, bm_set) in enumerate(zip(battles, battle_completed, bold_members), start=1):
        ticket_src = f"{b.ticket_source[0]}:{b.ticket_source[1]}" if b.ticket_source[0] else ""
        row = [i, b.target, b.ticket_kind, ticket_src]
        for m in members:
            row.append(b.participants.get(m, "—"))
        row.append(bc)
        ws.append(row)
        # Alternate target cell colors by map group (2 colors only).
        g = TARGET_TO_MAP.get(b.target, 0)
        ws.cell(row=ws.max_row, column=2).fill = (
            TARGET_GROUP_FILL_A if (g % 2 == 0) else TARGET_GROUP_FILL_B
        )
        # Bold the participant cells where this battle credits a quest.
        for j, m in enumerate(members, start=5):
            if m in bm_set:
                ws.cell(row=ws.max_row, column=j).font = BOLD_FONT
        if prev is not None:
            for j, m in enumerate(members, start=5):
                pc = prev.participants.get(m)
                nc = b.participants.get(m)
                if pc and nc and pc != nc:
                    ws.cell(row=ws.max_row, column=j).fill = SWITCH_FILL
        prev = b

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 10

    # --- Summary sheet ---
    ws2 = wb.create_sheet(SUMMARY_SHEET)
    ws2.append(["member", "quests_completed", "battles_participated",
                "distinct_characters_used", "character_switches"])
    for col in range(1, 6):
        c = ws2.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    per_member: Dict[str, Dict[str, object]] = {
        m: {"battles": 0, "chars": set(), "switches": 0} for m in members
    }
    prev = None
    for b in battles:
        for m, c in b.participants.items():
            per_member[m]["battles"] += 1
            per_member[m]["chars"].add(c)
            if prev is not None:
                pc = prev.participants.get(m)
                if pc is not None and pc != c:
                    per_member[m]["switches"] += 1
        prev = b

    for m in members:
        d = per_member[m]
        ws2.append([m, per_member_q[m] if quests is not None else "",
                    d["battles"], len(d["chars"]), d["switches"]])
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    total_battles = len(battles)
    total_quests = sum(battle_completed)
    ws2.append([])
    ws2.append(["TOTAL battles", total_battles])
    ws2.append(["TOTAL quests completed", total_quests])

    # Legend
    ws3 = wb.create_sheet("Legend")
    ws3.append(["Orange cell in Schedule = this member switches character vs previous battle."])
    ws3.append(["Bold character name = this battle credits that character's weekly quest."])
    ws3.append(["ticket_source = member:character who spends 1 ticket for that battle."])
    ws3.append(["— = member sits out (only possible for 双生, team size = 2)."])
    ws3.append(["quests_completed counts each (character, target) quest at most once per week."])
    ws3.column_dimensions["A"].width = 90

    wb.save(out_path)


def write_schedule_with_quests(
    battles: List[Battle],
    members: List[str],
    quests: Dict,              # (m,c,t) -> 0/1
    out_path: str | Path,
) -> None:
    """Write the schedule with correct once-per-week quest credit."""
    write_schedule(battles, members, out_path, quests=quests)
