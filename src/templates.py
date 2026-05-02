"""Generate one ticket + one quest template per team member.

For each member in MEMBERS we write two files:
    <member>_票.xlsx    — tickets (sheet: Characters)
    <member>_委托.xlsx  — weekly quests (sheet: Quests)

Column layout (no `ticket_` prefix; member is implied by filename):

  tickets:  character | 狮蝎 | 海龙 | ... | 双生 | 选择
  quests :  character | 狮蝎 | 海龙 | ... | 双生

Each file contains at least PLACEHOLDER_ROWS blank rows for the user to fill in.

Run:  python generate_templates.py [out_dir]
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

from .config import (
    CHAR_SHEET, QUEST_SHEET, TARGETS, WILDCARD, CHARACTER_COL,
    TICKET_SUFFIX, QUEST_SUFFIX, PLACEHOLDER_ROWS,
)

DEFAULT_TEMPLATE_MEMBERS = ["小C", "暗部", "桃核", "蹦蹦"]

HEADER_FILL = PatternFill("solid", fgColor="FFD9E1F2")
ROW_FILL_A = PatternFill("solid", fgColor="FFFFFFFF")
ROW_FILL_B = PatternFill("solid", fgColor="FFFFE6EA")
HEADER_FONT = Font(bold=True)
INFO_FONT = Font(italic=True, color="FF666666")
VERTICAL_BORDER = Border(
    left=Side(style="thin", color="FFA6A6A6"),
    right=Side(style="thin", color="FFA6A6A6"),
)


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = VERTICAL_BORDER
    ws.freeze_panes = "B2"
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 6
    ws.column_dimensions["A"].width = 12


def _add_placeholder_rows(ws, ncols: int, n_rows: int) -> None:
    """Emit blank fillable rows below the header.

    Numeric cells are intentionally blank; the input reader treats blanks as 0.
    Rows use alternating fills, and every cell gets vertical borders so columns
    are easy to follow while filling the workbook by hand.
    """
    for i in range(n_rows):
        row = [""] * ncols
        ws.append(row)
        fill = ROW_FILL_A if i % 2 == 0 else ROW_FILL_B
        for col in range(1, ncols + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.border = VERTICAL_BORDER
            cell.alignment = Alignment(horizontal="center")


def build_ticket_file(member: str, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = CHAR_SHEET

    headers = [CHARACTER_COL] + list(TARGETS) + [WILDCARD]
    ws.append(headers)
    _style_header(ws, len(headers))

    ws.cell(row=1, column=1).comment = Comment(
        f"This file belongs to member 「{member}」 (inferred from filename).\n"
        "Enter one row per character in your account. Character names must be unique.",
        "template")
    ws.cell(row=1, column=2).comment = Comment(
        "Weekly entry ticket count for this target (0, 1, 2, ...). "
        "A character may hold multiple tickets for the same target.",
        "template")
    ws.cell(row=1, column=len(headers)).comment = Comment(
        f"「{WILDCARD}」 = wildcard tickets, usable for ANY target except 双生.",
        "template")

    _add_placeholder_rows(ws, len(headers), PLACEHOLDER_ROWS)

    wb.save(path)


def build_quest_file(member: str, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = QUEST_SHEET

    headers = [CHARACTER_COL] + list(TARGETS)
    ws.append(headers)
    _style_header(ws, len(headers))

    ws.cell(row=1, column=1).comment = Comment(
        f"This file belongs to member 「{member}」 (inferred from filename).\n"
        "One row per character. Character names must match your ticket file.",
        "template")
    ws.cell(row=1, column=2).comment = Comment(
        "1 = this target is a weekly quest for the character; 0 = not a quest.",
        "template")

    _add_placeholder_rows(ws, len(headers), PLACEHOLDER_ROWS)

    wb.save(path)


def main(
    out_dir: str | Path = "templates",
    members: list[str] | None = None,
) -> None:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    members = members or DEFAULT_TEMPLATE_MEMBERS
    for member in members:
        build_ticket_file(member, out / f"{member}{TICKET_SUFFIX}.xlsx")
        build_quest_file(member, out / f"{member}{QUEST_SUFFIX}.xlsx")
    print(f"Wrote {len(members) * 2} template files to {out.resolve()}")


if __name__ == "__main__":
    main()
