"""Regression checks for dynamic members, blank numeric cells, and target coloring.

Run:
    python -m tests.test_dynamic_features
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.config import (
    CHAR_SHEET,
    QUEST_SHEET,
    TARGETS,
    WILDCARD,
    CHARACTER_COL,
    SCHEDULE_SHEET,
    TARGET_TO_MAP,
    TICKET_SUFFIX,
    QUEST_SUFFIX,
)
from src.optimize import _load_inputs, solve
from src.schedule import finalize_schedule, write_schedule_with_quests
from tests.validate_schedule import validate


def _ticket_row(name: str, overrides: dict[str, int], wildcard: int | str | None) -> dict:
    row = {CHARACTER_COL: name}
    for t in TARGETS:
        row[t] = "" if t not in overrides else overrides[t]
    row[WILDCARD] = wildcard
    return row


def _quest_row(name: str, overrides: dict[str, int]) -> dict:
    row = {CHARACTER_COL: name}
    for t in TARGETS:
        # Mix None and empty strings to validate coercion-to-zero behavior.
        row[t] = overrides[t] if t in overrides else (None if hash((name, t)) % 2 else "")
    return row


def main() -> None:
    root = Path(__file__).resolve().parent
    input_dir = root / "tmp_dynamic_input"
    output_file = root / "tmp_dynamic_output.xlsx"

    if input_dir.exists():
        shutil.rmtree(input_dir)
    input_dir.mkdir(parents=True, exist_ok=True)
    if output_file.exists():
        output_file.unlink()

    members = ["Alpha", "Bravo", "Charlie"]

    for m in members:
        c = f"{m}_C1"
        if m == "Alpha":
            ticket_rows = [_ticket_row(c, {"狮蝎": 2, "K博士": 1}, "")]
        elif m == "Bravo":
            ticket_rows = [_ticket_row(c, {"狮蝎": 1, "K博士": 1}, None)]
        else:
            ticket_rows = [_ticket_row(c, {"狮蝎": 1, "K博士": 1}, 1)]

        quest_rows = [_quest_row(c, {"狮蝎": 1, "K博士": 1})]

        pd.DataFrame(ticket_rows, columns=[CHARACTER_COL] + list(TARGETS) + [WILDCARD]).to_excel(
            input_dir / f"{m}{TICKET_SUFFIX}.xlsx",
            sheet_name=CHAR_SHEET,
            index=False,
        )
        pd.DataFrame(quest_rows, columns=[CHARACTER_COL] + list(TARGETS)).to_excel(
            input_dir / f"{m}{QUEST_SUFFIX}.xlsx",
            sheet_name=QUEST_SHEET,
            index=False,
        )

    loaded_members, chars_by_member, tickets, _, quests = _load_inputs(input_dir)
    assert loaded_members == sorted(members), f"dynamic member discovery failed: {loaded_members}"

    # Blank/None cells should be interpreted as zero.
    assert tickets[("Alpha", "Alpha_C1", "海龙")] == 0
    assert quests[("Alpha", "Alpha_C1", "海龙")] == 0

    battles, solved_members = solve(input_dir, time_limit_sec=30)
    _, chars_by_member, _, _, quests = _load_inputs(input_dir)
    battles = finalize_schedule(battles, solved_members, quests, chars_by_member)
    write_schedule_with_quests(battles, solved_members, quests, output_file)

    assert validate(input_dir, output_file), "validator failed for dynamic-member case"

    wb = load_workbook(output_file)
    ws = wb[SCHEDULE_SHEET]

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    for m in members:
        assert m in headers, f"missing dynamic member column: {m}"
    assert "小C" not in headers, "legacy fixed member leaked into schedule columns"

    # Target-cell fill color should alternate by contiguous map block in schedule order.
    color_a = "FFF2F2F2"
    color_b = "FFFFE6EA"
    seen = set()
    prev_map = None
    prev_fill = None
    for r in range(2, ws.max_row + 1):
        target = ws.cell(row=r, column=2).value
        fill_rgb = ws.cell(row=r, column=2).fill.fgColor.rgb
        current_map = TARGET_TO_MAP.get(target, -1)
        if prev_map is not None:
            if current_map == prev_map:
                assert fill_rgb == prev_fill, (
                    f"same-map target block changed color at row {r}: target={target}"
                )
            else:
                assert fill_rgb != prev_fill, (
                    f"map switch did not change color at row {r}: target={target}"
                )
        seen.add(fill_rgb)
        prev_map = current_map
        prev_fill = fill_rgb

    assert seen.issubset({color_a, color_b}), f"unexpected target fill colors: {seen}"

    print("[OK] dynamic members, blank-as-zero parsing, and target coloring all validated.")


if __name__ == "__main__":
    main()
