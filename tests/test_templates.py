"""Regression checks for generated template workbook styling.

Run:
    python -m tests.test_templates
"""
from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from src.config import CHAR_SHEET, QUEST_SHEET, TARGETS, WILDCARD, PLACEHOLDER_ROWS
from src.templates import main as generate_templates

ROW_FILL_WHITE = "FFFFFFFF"
ROW_FILL_PINK = "FFFFE6EA"
BORDER_GREY = "FFA6A6A6"


def _check_sheet(path: Path, sheet: str, ncols: int) -> None:
    wb = load_workbook(path)
    ws = wb[sheet]

    assert ws.max_row == PLACEHOLDER_ROWS + 1, (
        f"expected {PLACEHOLDER_ROWS} fillable rows in {path.name}, got {ws.max_row - 1}"
    )

    row2_values = [ws.cell(row=2, column=c).value for c in range(1, ncols + 1)]
    assert all(v is None for v in row2_values), f"expected blank fillable cells in {path.name}"

    for row in (1, 2, 3):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            assert cell.border.left.style == "thin", f"missing left border in {path.name} {cell.coordinate}"
            assert cell.border.right.style == "thin", f"missing right border in {path.name} {cell.coordinate}"
            assert cell.border.left.color.rgb == BORDER_GREY, f"wrong left border color in {path.name} {cell.coordinate}"
            assert cell.border.right.color.rgb == BORDER_GREY, f"wrong right border color in {path.name} {cell.coordinate}"

    assert ws.cell(row=2, column=1).fill.fgColor.rgb == ROW_FILL_WHITE, f"row 2 should be white in {path.name}"
    assert ws.cell(row=3, column=1).fill.fgColor.rgb == ROW_FILL_PINK, f"row 3 should be light pink in {path.name}"


def main() -> None:
    root = Path(__file__).resolve().parent
    out_dir = root / "tmp_templates"
    if out_dir.exists():
        shutil.rmtree(out_dir)

    generate_templates(out_dir, members=["Alpha", "Bravo"])

    ticket_cols = 1 + len(TARGETS) + 1  # 角色 + targets + 选择
    quest_cols = 1 + len(TARGETS)       # 角色 + targets
    for member in ("Alpha", "Bravo"):
        _check_sheet(out_dir / f"{member}_票.xlsx", CHAR_SHEET, ticket_cols)
        _check_sheet(out_dir / f"{member}_委托.xlsx", QUEST_SHEET, quest_cols)

    print("[OK] generated templates use blank cells, vertical borders, and alternating row colors.")


if __name__ == "__main__":
    main()
